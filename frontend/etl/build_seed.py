#!/usr/bin/env python3
"""
Load the company-provided Excel workbooks into the web client's data layer.

    python etl/build_seed.py

Reads the four ESCA_HSE_*.xlsx workbooks and writes
`web/src/api/mock/seed.generated.js` — a faithful mirror of the sheets with
foreign keys left intact.

Two rules this script follows:

1. It does not invent data. Every value in the output comes from a cell.
   Arabic labels for coded values (statuses, severities, permit types) are a
   presentation concern and live in `web/src/labels.js`, not here.

2. It is re-runnable. When a new drop of sheets arrives, re-run it — the
   generated file is overwritten and nothing else needs touching. That is why
   the output carries a header telling people not to edit it by hand.

Sheet quirks handled: each sheet starts with a title banner and a blank row
before the real header, so the header row is detected rather than assumed.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python -m pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_SOURCE = Path.home() / "Downloads"
OUT_FILE = ROOT / "web" / "src" / "api" / "mock" / "seed.generated.js"

WORKBOOKS = [
    "ESCA_HSE_01_Master_Data_and_Dictionary.xlsx",
    "ESCA_HSE_02_Core_Operations_Sample_Data.xlsx",
    "ESCA_HSE_03_Assets_Training_Health_Sample_Data.xlsx",
    "ESCA_HSE_04_AI_IoT_Integrations_Audit_Sample_Data.xlsx",
]

# sheet name -> exported constant name
EXPORTS = {
    # 01 — master data
    "Departments": "departments",
    "Zones": "zones",
    "Employees": "employees",
    "Roles": "roles",
    "Users": "users",
    "User_Roles": "userRoles",
    "RBAC_Matrix": "rbacMatrix",
    "Service_Catalog": "serviceCatalog",
    "Environments": "environments",
    # 02 — core operations
    "Incidents": "incidents",
    "Incident_RCA": "incidentRca",
    "CAPA": "capa",
    "Inspections": "inspections",
    "Findings": "findings",
    "Inspection_Responses": "inspectionResponses",
    "Risk_Register": "risks",
    "JSA": "jsa",
    "JSA_Steps": "jsaSteps",
    "Permits": "permits",
    "Permit_Approvals": "permitApprovals",
    "Permit_Checklist": "permitChecklist",
    "Permit_Gas_Tests": "permitGasTests",
    "SIMOPS": "simops",
    "Monthly_KPIs": "monthlyKpis",
    "Report_Definitions": "reportDefinitions",
    "Report_Runs": "reportRuns",
    # 03 — assets, training, health
    "Training_Courses": "trainingCourses",
    "Training_Requirements": "trainingRequirements",
    "Certificates": "certificates",
    "PPE_Inventory": "ppeInventory",
    "PPE_Matrix": "ppeMatrix",
    "PPE_Transactions": "ppeTransactions",
    "Fire_Equipment": "fireEquipment",
    "Fire_Inspections": "fireInspections",
    "Fixed_Safety_Assets": "fixedSafetyAssets",
    "Chemicals": "chemicals",
    "SDS_Records": "sdsRecords",
    "Medical_Protocols": "medicalProtocols",
    "Employee_Exposures": "employeeExposures",
    "Health_Exams": "healthExams",
    "Clinic_Visits": "clinicVisits",
    # 04 — AI, IoT, integrations, audit
    "IoT_Sensors": "iotSensors",
    "Sensor_Readings": "sensorReadings",
    "Cameras": "cameras",
    "AI_Models": "aiModels",
    "AI_Events": "aiEvents",
    "Wearable_Devices": "wearableDevices",
    "Wearable_Events": "wearableEvents",
    "Integrations": "integrations",
    "API_Logs": "apiLogs",
    "QA_Sessions": "qaSessions",
    "QA_Messages": "qaMessages",
    "QA_Tool_Calls": "qaToolCalls",
    "Notifications": "notifications",
    "Audit_Log": "auditLog",
    "Security_Events": "securityEvents",
    "Automation_Rules": "automationRules",
    "Automation_Runs": "automationRuns",
    "Automation_Actions": "automationActions",
}

SNAKE = re.compile(r"^[a-z][a-z0-9_]*$")


def cell(v):
    """Normalise a cell value into something JSON can carry."""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.isoformat()[:19]
    if isinstance(v, str):
        s = v.strip()
        return s or None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def find_header(rows: list[list]) -> int | None:
    """
    The real header is the first row that looks like database column names.

    Every sheet opens with a title banner ("ESCA HSE | Departments"), a
    description line and a blank row, so row 1 is never the header. Rather than
    hard-coding "row 4", detect it — a later drop of sheets may add a line.
    """
    for i, row in enumerate(rows):
        vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(vals) >= 2 and sum(bool(SNAKE.match(v)) for v in vals) / len(vals) >= 0.7:
            return i
    return None


def read_sheet(ws) -> list[dict]:
    rows = [[cell(c) for c in r] for r in ws.iter_rows(values_only=True)]
    h = find_header(rows)
    if h is None:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[h]]
    keep = [i for i, c in enumerate(header) if c]
    out = []
    for r in rows[h + 1 :]:
        if not any(i < len(r) and r[i] is not None for i in keep):
            continue  # blank spacer row
        out.append({header[i]: (r[i] if i < len(r) else None) for i in keep})
    return out


def read_summary(ws) -> dict:
    """
    The Summary sheet is a hand-laid dashboard, not a table: two label/value
    blocks side by side with a header row in the middle. Read it positionally.
    """
    rows = [[cell(c) for c in r] for r in ws.iter_rows(values_only=True)]
    flat = {str(r[0]): r[1] for r in rows if r and r[0] and len(r) > 1}

    kpis, headcount = [], []
    start = next((i for i, r in enumerate(rows) if r and r[0] == "KPI"), None)
    if start is not None:
        for r in rows[start + 1 :]:
            if len(r) > 1 and r[0] and r[1] is not None:
                kpis.append({"label": str(r[0]), "value": r[1]})
            if len(r) > 4 and r[3] and r[4] is not None:
                headcount.append({"department": str(r[3]), "headcount": r[4]})

    return {
        "title": str(rows[0][0]) if rows and rows[0] and rows[0][0] else "",
        "subtitle": str(rows[1][0]) if len(rows) > 1 and rows[1] and rows[1][0] else "",
        "asOfDate": str(flat.get("As-of date", ""))[:10],
        "asOfTimestamp": str(flat.get("As-of timestamp", "")).replace("T", " ")[:16],
        "kpis": kpis,
        "headcountByDepartment": headcount,
    }


def main() -> int:
    source = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_SOURCE

    missing = [w for w in WORKBOOKS if not (source / w).exists()]
    if missing:
        print(f"Workbooks not found in {source}:", file=sys.stderr)
        for m in missing:
            print(f"  - {m}", file=sys.stderr)
        print("\nPass the folder holding them:  python etl/build_seed.py <folder>", file=sys.stderr)
        return 1

    tables: dict[str, list[dict]] = {}
    summary: dict = {}
    report: list[tuple[str, int]] = []

    for name in WORKBOOKS:
        wb = openpyxl.load_workbook(source / name, data_only=True)
        for ws in wb.worksheets:
            if ws.title == "Summary" and not summary:
                summary = read_summary(ws)
            const = EXPORTS.get(ws.title)
            if not const:
                continue
            rows = read_sheet(ws)
            tables[const] = rows
            report.append((ws.title, len(rows)))
        wb.close()

    unmapped = [s for s in EXPORTS.values() if s not in tables]
    if unmapped:
        print("WARNING — expected sheets that produced no rows:", ", ".join(unmapped), file=sys.stderr)

    parts = [
        "/* eslint-disable */",
        "/**",
        " * AUTO-GENERATED — do not edit by hand.",
        " *",
        " * Source: the four ESCA_HSE_*.xlsx workbooks provided by the plant.",
        " * Regenerate with:  python etl/build_seed.py [folder-with-workbooks]",
        " *",
        " * Field names and coded values are exactly as they appear in the sheets, so",
        " * this file doubles as the reference payload for the Spring Boot team.",
        " * Arabic labels for coded values live in src/labels.js, not here.",
        " */",
        "",
        f"export const meta = {json.dumps({'generatedFrom': WORKBOOKS, 'sheetCount': len(tables), 'rowCount': sum(len(v) for v in tables.values())}, ensure_ascii=False, indent=2)}",
        "",
        f"export const summary = {json.dumps(summary, ensure_ascii=False, indent=2)}",
        "",
    ]

    for const in EXPORTS.values():
        rows = tables.get(const, [])
        parts.append(f"export const {const} = {json.dumps(rows, ensure_ascii=False, indent=1)}")
        parts.append("")

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text("\n".join(parts), encoding="utf-8")

    total = sum(n for _, n in report)
    print(f"Wrote {OUT_FILE.relative_to(ROOT)}")
    print(f"  {len(report)} sheets, {total} rows")
    for sheet, n in report:
        print(f"    {sheet:<24} {n:>5}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
